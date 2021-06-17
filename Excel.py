import openpyxl
from DataStructures import DoublyLinkedList


# Function to import data from excel file
def import_excel_data(off_peak_selected):  # with excel changes
    data_book = openpyxl.load_workbook('London Underground data.xlsx')
    sheet = data_book.active
    dll = DoublyLinkedList()  # Initialising doubly linked list
    # Upload station names as nodes of the Doubly Linked List
    for row in sheet.iter_rows(min_row=2, min_col=2, max_row=759, max_col=2):
        for cell in row:
            repeated = False
            if cell.value is not None:
                for node in dll.iterate():
                    if cell.value == node[0]:
                        repeated = True
                if not repeated:
                    dll.append(cell.value)

    # Upload details of each station
    # row = (Line, Station, Neighbour, Distance in minutes)
    for node in dll.iterate():
        for row in sheet.iter_rows(min_row=26, min_col=1, max_row=759, max_col=4):
            if row[2].value is not None:
                if row[1].value == node[0]:
                    if off_peak_selected and row[0].value == 'Bakerloo':
                        time = row[3].value / 2
                        dll.search_node(row[1].value, row[2].value, time, row[0].value)
                        dll.search_node(row[2].value, row[1].value, time, row[0].value)
                    else:
                        dll.search_node(row[1].value, row[2].value, row[3].value, row[0].value)
                        dll.search_node(row[2].value, row[1].value, row[3].value, row[0].value)

    return dll
